"""Export 模块 - Excel 导出与 Gantt 图

功能:
- 导出时延分析结果到 xlsx
- 生成流水图 (Gantt Chart) 页签
- 支持多种格式化选项
"""

from pathlib import Path

from .latency import GanttData, LatencyResult

# Gantt 图颜色映射
GANTT_COLORS = {
    "cube": "4472C4",  # 蓝色
    "vector": "70AD47",  # 绿色
    "dma": "FFC000",  # 黄色
    "execution": "4472C4",  # 蓝色
    "prefetch": "FFC000",  # 黄色
    "parallel": "7030A0",  # 紫色
}


def export_xlsx(
    result: LatencyResult,
    output_path: str,
    include_gantt: bool = True,
    include_passes: bool = True,
    include_summary: bool = True,
):
    """导出分析结果到 xlsx

    Args:
        result: 分析结果
        output_path: 输出文件路径
        include_gantt: 是否包含 Gantt 图页签
        include_passes: 是否包含 Pass 详情页签
        include_summary: 是否包含摘要页签
    """
    try:
        import openpyxl
        from openpyxl.styles import Border, Font, PatternFill, Side
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for xlsx export. Install with: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 1. 主分析结果页签
    ws_main = wb.active
    ws_main.title = "Latency Analysis"
    _write_main_sheet(ws_main, result, header_font, header_fill, border)

    # 2. 摘要页签
    if include_summary and result.summary:
        ws_summary = wb.create_sheet("Summary")
        _write_summary_sheet(ws_summary, result, header_font, header_fill, border)

    # 3. Pass 详情页签
    if include_passes and result.pass_results:
        ws_passes = wb.create_sheet("Pass Details")
        _write_passes_sheet(ws_passes, result, header_font, header_fill, border)

    # 4. Pass 配置页签
    if include_passes and result.pass_config:
        ws_config = wb.create_sheet("Pass Config")
        _write_config_sheet(ws_config, result.pass_config, header_font, header_fill, border)

    # 5. 计算详情页签 (可手动重算验证)
    ws_calc = wb.create_sheet("Calculation Details")
    _write_calculation_sheet(ws_calc, result, header_font, header_fill, border)

    # 6. Gantt 图页签
    if include_gantt and result.gantt_data:
        ws_gantt = wb.create_sheet("Gantt Chart")
        _write_gantt_sheet(ws_gantt, result.gantt_data, header_font, header_fill, border)

    # 保存文件
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Exported to: {output_path}")


def _write_main_sheet(ws, result: LatencyResult, header_font, header_fill, border):
    """写入主分析结果页签"""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    # 表头
    headers = [
        "Op Name",
        "Op Type",
        "Compute Unit",
        "Dtype",
        "FLOPs (M)",
        "Input (KB)",
        "Weight (KB)",
        "Output (KB)",
        "Compute (us)",
        "Memory (us)",
        "Roofline (us)",
        "Prefetch Saved",
        "Parallel Saved",
        "Overhead",
        "Total (us)",
        "Bottleneck",
        "Min BW (GB/s)",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for row_idx, bd in enumerate(result.breakdowns, 2):
        p = bd.profile
        values = [
            p.name,
            p.op_type,
            p.compute_unit,
            p.dtype,
            p.flops / 1e6,
            p.input_bytes / 1024,
            p.weight_bytes / 1024,
            p.output_bytes / 1024,
            bd.timing.compute_us,
            bd.timing.memory_us,
            bd.timing.roofline_us,
            bd.savings.prefetch_us,
            bd.savings.parallel_us,
            bd.timing.overhead_us,
            bd.timing.total_us,
            bd.bottleneck,
            bd.bandwidth.min_gbps,
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if isinstance(value, float):
                cell.number_format = "0.00"

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _write_summary_sheet(ws, result: LatencyResult, _header_font, _header_fill, _border):
    """写入摘要页签"""
    from openpyxl.styles import Font

    s = result.summary
    chip = result.chip_spec

    # 标题
    ws.cell(row=1, column=1, value=f"Paper Analysis Summary - {chip.name}").font = Font(
        bold=True, size=14
    )
    ws.merge_cells("A1:D1")

    # 基本信息
    data = [
        ("", ""),
        ("Chip Info", ""),
        ("Chip Name", chip.name),
        ("Cube FP16 TFLOPS", chip.cube.fp16_tflops),
        ("Vector FP16 GFLOPS", chip.vector.fp16_gflops),
        ("HBM Bandwidth (GB/s)", chip.memory.hbm.bandwidth_gbps),
        ("HBM Capacity (GB)", chip.memory.hbm.capacity_bytes / 1024**3),
        ("", ""),
        ("Analysis Summary", ""),
        ("Total Operators", len(result.breakdowns)),
        ("Total Latency (us)", s.totals.latency_us),
        ("Total Latency (ms)", s.totals.latency_us / 1000),
        ("", ""),
        ("Time Breakdown", ""),
        ("Compute Time (us)", s.totals.compute_time_us),
        ("Memory Time (us)", s.totals.memory_time_us),
        ("Overhead (us)", s.optimization.overhead_us),
        ("", ""),
        ("Bottleneck Stats", ""),
        ("Compute Bound Ops", s.bottleneck.compute_bound_ops),
        ("Memory Bound Ops", s.bottleneck.memory_bound_ops),
        ("", ""),
        ("Optimizations", ""),
        ("Prefetch Saved (us)", s.optimization.prefetch_saved_us),
        ("Parallel Saved (us)", s.optimization.parallel_saved_us),
        ("", ""),
        ("Throughput", ""),
        ("Achieved TFLOPS", s.throughput.achieved_tflops),
        ("Achieved Bandwidth (GB/s)", s.throughput.achieved_bandwidth_gbps),
        ("", ""),
        ("Unit Utilization", ""),
        ("Cube Time (us)", s.unit.cube_time_us),
        ("Vector Time (us)", s.unit.vector_time_us),
        (
            "Cube Ratio (%)",
            s.unit.cube_time_us / s.totals.latency_us * 100 if s.totals.latency_us > 0 else 0,
        ),
        (
            "Vector Ratio (%)",
            s.unit.vector_time_us / s.totals.latency_us * 100 if s.totals.latency_us > 0 else 0,
        ),
    ]

    for row_idx, (label, value) in enumerate(data, 3):
        ws.cell(row=row_idx, column=1, value=label)
        if value != "":
            cell = ws.cell(row=row_idx, column=2, value=value)
            if isinstance(value, float):
                cell.number_format = "0.00"

        # 小节标题加粗
        if label and value == "":
            ws.cell(row=row_idx, column=1).font = Font(bold=True)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15


def _write_passes_sheet(ws, result: LatencyResult, header_font, header_fill, border):
    """写入 Pass 详情页签"""
    from openpyxl.utils import get_column_letter

    # 表头
    headers = [
        "Op Name",
        "Pass Name",
        "Enabled",
        "Before (us)",
        "After (us)",
        "Saved (us)",
        "Improvement %",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    row_idx = 2
    for bd, passes in zip(result.breakdowns, result.pass_results):
        for pr in passes:
            values = [
                bd.profile.name,
                pr.pass_name,
                "Yes" if pr.enabled else "No",
                pr.latency_before_us,
                pr.latency_after_us,
                pr.latency_saved_us,
                pr.improvement_ratio * 100 if pr.latency_before_us > 0 else 0,
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if isinstance(value, float):
                    cell.number_format = "0.00"

            row_idx += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _write_config_sheet(ws, pass_config, header_font, header_fill, border):
    """写入 Pass 配置页签"""
    from openpyxl.styles import Alignment, Font, PatternFill

    # 标题
    ws.cell(row=1, column=1, value="Pass Configuration").font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    # 预设信息
    ws.cell(row=3, column=1, value="Preset").font = Font(bold=True)
    ws.cell(row=3, column=2, value=pass_config.preset.value)

    # Pass 启用状态表
    row = 5
    ws.cell(row=row, column=1, value="Pass Name").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="Enabled").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=3, value="Key Parameters").font = header_font
    ws.cell(row=row, column=3).fill = header_fill

    # Pass 配置列表
    passes_info = [
        ("Roofline", pass_config.roofline_enabled, "-"),
        (
            "MinTraffic",
            pass_config.min_traffic.enabled,
            f"l2_reuse={pass_config.min_traffic.l2_reuse_factor}, "
            f"tiling_eff={pass_config.min_traffic.tiling_efficiency}",
        ),
        (
            "MemoryEfficiency",
            pass_config.memory_efficiency_enabled,
            f"use_effective_bw={pass_config.use_effective_bandwidth}",
        ),
        (
            "BandwidthConstraint",
            pass_config.bandwidth.enabled,
            f"streams={pass_config.bandwidth.concurrent_streams}, "
            f"model={pass_config.bandwidth.contention_model}",
        ),
        (
            "ForwardPrefetch",
            pass_config.prefetch.forward_enabled,
            f"efficiency={pass_config.prefetch.efficiency}",
        ),
        (
            "BackwardPrefetch",
            pass_config.prefetch.backward_enabled,
            f"depth={pass_config.prefetch.backward_depth}",
        ),
        ("CubeVectorParallel", pass_config.cube_vector_parallel_enabled, "-"),
        (
            "Overhead",
            pass_config.overhead.enabled,
            f"kernel={pass_config.overhead.kernel_launch_us}us, "
            f"sync={pass_config.overhead.sync_us}us, "
            f"ctx={pass_config.overhead.context_switch_us}us, "
            f"tiling={pass_config.overhead.tiling_us}us",
        ),
        (
            "TrafficConstraint",
            pass_config.traffic.enabled,
            f"max={pass_config.traffic.max_bytes}, mode={pass_config.traffic.budget_mode}",
        ),
    ]

    # 颜色
    enabled_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    disabled_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for pass_name, enabled, params in passes_info:
        row += 1
        ws.cell(row=row, column=1, value=pass_name).border = border
        cell = ws.cell(row=row, column=2, value="Yes" if enabled else "No")
        cell.fill = enabled_fill if enabled else disabled_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=params).border = border

    # 详细参数表
    row += 3
    ws.cell(row=row, column=1, value="All Parameters").font = Font(bold=True, size=12)

    row += 1
    ws.cell(row=row, column=1, value="Parameter").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="Value").font = header_font
    ws.cell(row=row, column=2).fill = header_fill

    params = [
        ("enabled", pass_config.enabled),
        ("preset", pass_config.preset.value),
        ("roofline_enabled", pass_config.roofline_enabled),
        ("memory_efficiency_enabled", pass_config.memory_efficiency_enabled),
        ("use_effective_bandwidth", pass_config.use_effective_bandwidth),
        ("prefetch.forward_enabled", pass_config.prefetch.forward_enabled),
        ("prefetch.efficiency", pass_config.prefetch.efficiency),
        ("prefetch.backward_enabled", pass_config.prefetch.backward_enabled),
        ("prefetch.backward_depth", pass_config.prefetch.backward_depth),
        ("cube_vector_parallel_enabled", pass_config.cube_vector_parallel_enabled),
        ("overhead.enabled", pass_config.overhead.enabled),
        ("overhead.kernel_launch_us", pass_config.overhead.kernel_launch_us),
        ("overhead.sync_us", pass_config.overhead.sync_us),
        ("overhead.context_switch_us", pass_config.overhead.context_switch_us),
        ("overhead.tiling_us", pass_config.overhead.tiling_us),
        ("overhead.tiling_count", pass_config.overhead.tiling_count),
        ("bandwidth.enabled", pass_config.bandwidth.enabled),
        ("bandwidth.concurrent_streams", pass_config.bandwidth.concurrent_streams),
        ("bandwidth.contention_model", pass_config.bandwidth.contention_model),
        ("traffic.enabled", pass_config.traffic.enabled),
        ("traffic.max_bytes", pass_config.traffic.max_bytes),
        ("traffic.budget_mode", pass_config.traffic.budget_mode),
        ("min_traffic.enabled", pass_config.min_traffic.enabled),
        ("min_traffic.cache_line_bytes", pass_config.min_traffic.cache_line_bytes),
        ("min_traffic.l2_reuse_factor", pass_config.min_traffic.l2_reuse_factor),
        ("min_traffic.tiling_efficiency", pass_config.min_traffic.tiling_efficiency),
    ]

    for param_name, param_value in params:
        row += 1
        ws.cell(row=row, column=1, value=param_name).border = border
        cell = ws.cell(row=row, column=2, value=str(param_value))
        cell.border = border
        if isinstance(param_value, float):
            cell.number_format = "0.00"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 50


def _write_chip_params_section(ws, chip, border, start_row: int) -> int:
    """写入芯片参数部分"""
    from openpyxl.styles import Font

    ws.cell(row=start_row, column=1, value="Chip Parameters (用于计算)").font = Font(
        bold=True, size=14
    )
    ws.merge_cells(f"A{start_row}:D{start_row}")

    chip_params = [
        ("Chip Name", chip.name, "", ""),
        (
            "Cube FP16 TFLOPS",
            chip.cube.fp16_tflops,
            "TFLOPS",
            "Cube 计算时延 = FLOPs / (TFLOPS × 1e12) × 1e6 us",
        ),
        (
            "Vector FP16 GFLOPS",
            chip.vector.fp16_gflops,
            "GFLOPS",
            "Vector 计算时延 = FLOPs / (GFLOPS × 1e9) × 1e6 us",
        ),
        (
            "HBM Bandwidth",
            chip.memory.hbm.bandwidth_gbps,
            "GB/s",
            "访存时延 = Bytes / (BW × 1e9) × 1e6 us",
        ),
    ]

    row = start_row + 2
    for label, value, unit, formula in chip_params:
        ws.cell(row=row, column=1, value=label).border = border
        cell = ws.cell(row=row, column=2, value=value)
        cell.border = border
        if isinstance(value, float):
            cell.number_format = "0.00"
        ws.cell(row=row, column=3, value=unit).border = border
        ws.cell(row=row, column=4, value=formula).font = Font(italic=True, color="666666")
        row += 1

    return row


def _write_operator_details_section(
    ws, result: LatencyResult, header_font, header_fill, border, start_row: int
) -> int:
    """写入算子详情部分"""
    from openpyxl.styles import Alignment, Font

    from .profile import dtype_bytes

    row = start_row + 2
    ws.cell(row=row, column=1, value="Operator Calculation Details").font = Font(bold=True, size=14)
    ws.merge_cells(f"A{row}:R{row}")
    row += 2

    headers = [
        "Op Name",
        "Op Type",
        "Unit",
        "Dtype",
        "Bytes/Elem",
        "M",
        "N",
        "K",
        "Input (B)",
        "Weight (B)",
        "Output (B)",
        "Total (B)",
        "FLOPs",
        "AI (FLOPs/B)",
        "Compute (us)",
        "Memory (us)",
        "Roofline (us)",
        "Bottleneck",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row += 1

    for bd in result.breakdowns:
        p = bd.profile
        shapes = p.shapes or {}

        m = shapes.get("M") or shapes.get("m") or shapes.get("batch", 1) * shapes.get("seq_len", 1)
        n = shapes.get("N") or shapes.get("n") or shapes.get("out_features", 0)
        k = shapes.get("K") or shapes.get("k") or shapes.get("in_features", 0)

        elem_bytes = dtype_bytes(p.dtype)
        m = m if m else "-"
        n = n if n else "-"
        k = k if k else "-"

        values = [
            p.name,
            p.op_type,
            p.compute_unit,
            p.dtype,
            elem_bytes,
            m,
            n,
            k,
            p.input_bytes,
            p.weight_bytes,
            p.output_bytes,
            p.total_bytes,
            p.flops,
            p.arithmetic_intensity,
            bd.timing.compute_us,
            bd.timing.memory_us,
            bd.timing.roofline_us,
            bd.bottleneck,
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if isinstance(value, float):
                cell.number_format = "0.0" if col == 14 else "0.00"

        row += 1

    return row


def _write_formulas_section(ws, start_row: int) -> int:
    """写入公式说明部分"""
    from openpyxl.styles import Font

    row = start_row + 2
    ws.cell(row=row, column=1, value="Calculation Formulas (计算公式)").font = Font(
        bold=True, size=14
    )
    ws.merge_cells(f"A{row}:F{row}")
    row += 2

    formulas = [
        ("MatMul FLOPs", "2 × M × N × K", "矩阵乘法浮点运算次数"),
        ("Conv2D FLOPs", "2 × N × H × W × C × K × R × S", "卷积浮点运算次数"),
        ("Input Bytes", "M × K × bytes_per_elem", "输入张量访存量"),
        ("Weight Bytes", "K × N × bytes_per_elem", "权重张量访存量"),
        ("Output Bytes", "M × N × bytes_per_elem", "输出张量访存量"),
        ("Arithmetic Intensity", "FLOPs / Total_Bytes", "计算访存比"),
        ("Compute Time (Cube)", "FLOPs / (Cube_TFLOPS × 1e12) × 1e6", "Cube 计算时延 (us)"),
        ("Compute Time (Vector)", "FLOPs / (Vector_GFLOPS × 1e9) × 1e6", "Vector 计算时延 (us)"),
        ("Memory Time", "Total_Bytes / (HBM_BW × 1e9) × 1e6", "访存时延 (us)"),
        ("Roofline Time", "max(Compute_Time, Memory_Time)", "Roofline 时延"),
        ("Bottleneck", "compute if Compute > Memory else memory", "瓶颈判断"),
        ("", "", ""),
        ("--- Overhead ---", "", "开销计算"),
        ("Kernel Launch", "kernel_launch_us", "kernel 启动开销"),
        ("Sync Overhead", "sync_overhead_us", "同步开销"),
        ("Context Switch", "context_switch_us", "算子切换时延"),
        ("Tiling Overhead", "tiling_overhead_us × tiling_count", "Tiling 调度开销"),
        ("Total Overhead", "kernel + sync + ctx_switch + tiling", "总开销"),
        ("Final Latency", "roofline + overhead - prefetch - parallel", "最终时延"),
    ]

    for name, formula, desc in formulas:
        ws.cell(row=row, column=1, value=name).font = Font(bold=True)
        ws.cell(row=row, column=2, value=formula).font = Font(name="Consolas")
        ws.cell(row=row, column=4, value=desc).font = Font(italic=True, color="666666")
        row += 1

    return row


def _write_pass_effects_section(
    ws, result: LatencyResult, header_font, header_fill, border, start_row: int
) -> int:
    """写入 Pass 优化效果部分"""
    from openpyxl.styles import Font

    row = start_row + 2
    ws.cell(row=row, column=1, value="Pass Effects (优化效果)").font = Font(bold=True, size=14)
    ws.merge_cells(f"A{row}:H{row}")
    row += 2

    pass_headers = [
        "Op Name",
        "Prefetch Saved",
        "Backward Prefetch",
        "Parallel Saved",
        "Overhead",
        "Total Time",
        "vs Roofline",
    ]
    for col, header in enumerate(pass_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    row += 1
    for bd in result.breakdowns:
        delta = bd.timing.total_us - bd.timing.roofline_us
        values = [
            bd.profile.name,
            bd.savings.prefetch_us,
            bd.savings.backward_prefetch_us,
            bd.savings.parallel_us,
            bd.timing.overhead_us,
            bd.timing.total_us,
            f"{delta:+.2f}" if delta != 0 else "0",
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if isinstance(value, float):
                cell.number_format = "0.00"

        row += 1

    return row


def _write_calculation_sheet(ws, result: LatencyResult, header_font, header_fill, border):
    """写入计算详情页签 - 支持手动重算验证"""
    from openpyxl.utils import get_column_letter

    chip = result.chip_spec

    row = _write_chip_params_section(ws, chip, border, 1)
    row = _write_operator_details_section(ws, result, header_font, header_fill, border, row)
    row = _write_formulas_section(ws, row)
    _write_pass_effects_section(ws, result, header_font, header_fill, border, row)

    # 调整列宽
    col_widths = [18, 10, 8, 8, 10, 6, 6, 6, 12, 12, 12, 12, 14, 10, 12, 12, 12, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_gantt_sheet(ws, gantt_data: GanttData, _header_font, _header_fill, _border):
    """写入 Gantt 图页签

    使用条件格式和单元格着色模拟 Gantt 图
    """
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    # 标题
    ws.cell(row=1, column=1, value=f"Pipeline Gantt Chart - {gantt_data.chip_name}").font = Font(
        bold=True, size=14
    )
    ws.merge_cells("A1:Z1")

    # 时间刻度 (每列代表一定时间)
    time_scale_us = (
        gantt_data.total_duration_us / 50 if gantt_data.total_duration_us > 0 else 1
    )  # 50 列
    ws.cell(row=2, column=1, value=f"Time Scale: {time_scale_us:.2f} us/column")

    # 图例
    row = 4
    ws.cell(row=row, column=1, value="Legend:")
    row += 1

    for unit, color in [("Cube", "4472C4"), ("Vector", "70AD47"), ("DMA/Prefetch", "FFC000")]:
        ws.cell(row=row, column=1, value=unit)
        cell = ws.cell(row=row, column=2, value="")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # Gantt 表头
    row += 2
    ws.cell(row=row, column=1, value="Op Name").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Unit").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Start (us)").font = Font(bold=True)
    ws.cell(row=row, column=4, value="End (us)").font = Font(bold=True)
    ws.cell(row=row, column=5, value="Duration (us)").font = Font(bold=True)

    # 时间轴标题 (从第6列开始)
    for i in range(50):
        col = 6 + i
        time_val = i * time_scale_us
        if i % 10 == 0:
            ws.cell(row=row, column=col, value=f"{time_val:.0f}")
        ws.column_dimensions[get_column_letter(col)].width = 2

    row += 1

    # Gantt 条目
    for item in gantt_data.items:
        ws.cell(row=row, column=1, value=item.op_name)
        ws.cell(row=row, column=2, value=item.unit)
        ws.cell(row=row, column=3, value=item.start_us).number_format = "0.00"
        ws.cell(row=row, column=4, value=item.end_us).number_format = "0.00"
        ws.cell(row=row, column=5, value=item.end_us - item.start_us).number_format = "0.00"

        # 绘制 Gantt 条
        start_col = int(item.start_us / time_scale_us) + 6 if time_scale_us > 0 else 6
        end_col = int(item.end_us / time_scale_us) + 6 if time_scale_us > 0 else 6

        # 选择颜色
        color = GANTT_COLORS.get(item.unit, GANTT_COLORS.get(item.category, "808080"))

        for col in range(start_col, min(end_col + 1, 56)):  # 最多50列
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        row += 1

    # 总时间
    row += 1
    ws.cell(row=row, column=1, value="Total Time (us):").font = Font(bold=True)
    ws.cell(row=row, column=2, value=gantt_data.total_duration_us).number_format = "0.00"

    # 调整列宽
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12


def export_csv(result: LatencyResult, output_path: str):
    """导出为 CSV 格式"""
    import csv

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # 表头
        headers = [
            "Op Name",
            "Op Type",
            "Compute Unit",
            "Dtype",
            "FLOPs",
            "Input Bytes",
            "Weight Bytes",
            "Output Bytes",
            "Compute Time (us)",
            "Memory Time (us)",
            "Roofline Time (us)",
            "Prefetch Saved (us)",
            "Parallel Saved (us)",
            "Overhead (us)",
            "Total Time (us)",
            "Bottleneck",
            "Min Bandwidth (GB/s)",
        ]
        writer.writerow(headers)

        # 数据行
        for bd in result.breakdowns:
            p = bd.profile
            row = [
                p.name,
                p.op_type,
                p.compute_unit,
                p.dtype,
                p.flops,
                p.input_bytes,
                p.weight_bytes,
                p.output_bytes,
                bd.timing.compute_us,
                bd.timing.memory_us,
                bd.timing.roofline_us,
                bd.savings.prefetch_us,
                bd.savings.parallel_us,
                bd.timing.overhead_us,
                bd.timing.total_us,
                bd.bottleneck,
                bd.bandwidth.min_gbps,
            ]
            writer.writerow(row)

    print(f"Exported to: {output_path}")


def export_json(result: LatencyResult, output_path: str):
    """导出为 JSON 格式"""
    import json

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 构建可序列化的数据结构
    data = {
        "chip": {
            "name": result.chip_spec.name,
            "cube_fp16_tflops": result.chip_spec.cube.fp16_tflops,
            "vector_fp16_gflops": result.chip_spec.vector.fp16_gflops,
            "hbm_bandwidth_gbps": result.chip_spec.memory.hbm.bandwidth_gbps,
        },
        "summary": {
            "total_latency_us": result.summary.totals.latency_us,
            "total_flops": result.summary.totals.flops,
            "total_bytes": result.summary.totals.bytes,
            "compute_bound_ops": result.summary.bottleneck.compute_bound_ops,
            "memory_bound_ops": result.summary.bottleneck.memory_bound_ops,
            "achieved_tflops": result.summary.throughput.achieved_tflops,
            "achieved_bandwidth_gbps": result.summary.throughput.achieved_bandwidth_gbps,
        },
        "breakdowns": [],
    }

    for bd in result.breakdowns:
        p = bd.profile
        breakdown_data = {
            "op_name": p.name,
            "op_type": p.op_type,
            "compute_unit": p.compute_unit,
            "dtype": p.dtype,
            "flops": p.flops,
            "input_bytes": p.input_bytes,
            "weight_bytes": p.weight_bytes,
            "output_bytes": p.output_bytes,
            "compute_time_us": bd.timing.compute_us,
            "memory_time_us": bd.timing.memory_us,
            "roofline_time_us": bd.timing.roofline_us,
            "total_time_us": bd.timing.total_us,
            "bottleneck": bd.bottleneck,
            "min_bandwidth_gbps": bd.bandwidth.min_gbps,
        }
        data["breakdowns"].append(breakdown_data)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Exported to: {output_path}")
