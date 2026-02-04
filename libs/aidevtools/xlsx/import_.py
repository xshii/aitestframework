"""xlsx 导入

从 xlsx 解析配置并生成 Python 代码。
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from aidevtools.core.log import logger
from aidevtools.core.utils import parse_list, parse_shape


def _check_openpyxl():
    if not HAS_OPENPYXL:
        raise ImportError("xlsx 功能需要 openpyxl，请安装: pip install openpyxl")


def _gen_linear(lines, indent, cfg, shape, dtype, depends, input_var):
    """生成 linear 算子代码"""
    in_dim = shape[-1] if shape else 64
    if not depends:
        lines.append(f"{indent}w_{cfg.id} = np.random.randn({in_dim}, 256).astype(np.{dtype})")
    else:
        lines.append(
            f"{indent}w_{cfg.id} = np.random.randn({input_var}.shape[-1], 256).astype(np.{dtype})"
        )
    lines.append(f"{indent}out_{cfg.id} = F.linear({input_var}, w_{cfg.id})")


def _gen_matmul(lines, indent, cfg, shape, dtype, depends, input_var):
    """生成 matmul 算子代码"""
    if len(depends) >= 2:
        keys = list(depends.keys())
        lines.append(f"{indent}out_{cfg.id} = F.matmul({keys[0]}_{cfg.id}, {keys[1]}_{cfg.id})")
    else:
        dim = shape[-1] if shape else 64
        if not depends:
            lines.append(f"{indent}b_{cfg.id} = np.random.randn({dim}, {dim}).astype(np.{dtype})")
        else:
            d = f"{input_var}.shape[-1]"
            lines.append(f"{indent}b_{cfg.id} = np.random.randn({d}, {d}).astype(np.{dtype})")
        lines.append(f"{indent}out_{cfg.id} = F.matmul({input_var}, b_{cfg.id})")


def _gen_attention(lines, indent, cfg, depends, input_var):
    """生成 attention 算子代码"""
    if "q" in depends and "k" in depends and "v" in depends:
        lines.append(f"{indent}out_{cfg.id} = F.attention(q_{cfg.id}, k_{cfg.id}, v_{cfg.id})")
    else:
        lines.append(f"{indent}# attention 需要 q, k, v 三个输入")
        lines.append(f"{indent}out_{cfg.id} = {input_var}  # placeholder")


def _gen_binary_op(lines, indent, cfg, op_name, depends, input_var):
    """生成二元算子代码 (add, mul)"""
    if len(depends) >= 2:
        keys = list(depends.keys())
        lines.append(f"{indent}out_{cfg.id} = F.{op_name}({keys[0]}_{cfg.id}, {keys[1]}_{cfg.id})")
    else:
        lines.append(f"{indent}out_{cfg.id} = F.{op_name}({input_var}, {input_var})")


def _gen_unary_op(lines, indent, cfg, op_name, input_var):
    """生成一元算子代码 (relu, softmax 等)"""
    lines.append(f"{indent}out_{cfg.id} = F.{op_name}({input_var})")


def _generate_op_call(lines, indent, cfg, op_name, shape, dtype, depends, input_var):
    """生成算子调用代码 - 使用字典分发替代 if-elif 链"""
    # 一元算子
    unary_ops = {"relu", "softmax", "gelu", "sigmoid", "tanh", "silu", "layernorm", "rmsNorm"}
    # 二元算子
    binary_ops = {"add", "mul", "div"}

    if op_name == "linear":
        _gen_linear(lines, indent, cfg, shape, dtype, depends, input_var)
    elif op_name == "matmul":
        _gen_matmul(lines, indent, cfg, shape, dtype, depends, input_var)
    elif op_name == "attention":
        _gen_attention(lines, indent, cfg, depends, input_var)
    elif op_name in binary_ops:
        _gen_binary_op(lines, indent, cfg, op_name, depends, input_var)
    elif op_name in unary_ops:
        _gen_unary_op(lines, indent, cfg, op_name, input_var)
    else:
        lines.append(f"{indent}# 未知算子: {op_name}")
        lines.append(f"{indent}out_{cfg.id} = {input_var}  # placeholder")


@dataclass
class BinaryPaths:
    """二进制文件路径配置"""

    golden: str = ""  # golden 文件路径（留空=自动生成）
    result: str = ""  # result 文件路径
    input: str = ""  # input 文件路径
    weight: str = ""  # weight 文件路径
    sim_cmd: str = ""  # 仿真命令，支持占位符


@dataclass
class OpConfig:
    """算子配置 (使用组合模式)"""

    id: int
    op_name: str
    shape: Tuple[int, ...]
    dtype: str
    depends: str  # 原始依赖字符串
    qtype: str
    skip: bool
    note: str
    paths: BinaryPaths = field(default_factory=BinaryPaths)

    def parse_depends(self) -> Dict[str, List[int]]:
        """
        解析依赖关系

        Returns:
            依赖映射，格式: {"input_name": [row_ids]}

        示例:
            "" -> {}  (无依赖，随机输入)
            "0" -> {"x": [0]}  (单依赖)
            "1,2" -> {"a": [1], "b": [2]}  (双输入依赖)
            "q:0,k:1,v:2" -> {"q": [0], "k": [1], "v": [2]}  (命名依赖)
        """
        if not self.depends or self.depends.strip() == "":
            return {}

        depends_str = self.depends.strip()
        result = {}

        # 检查是否是命名依赖 (包含 ":")
        if ":" in depends_str:
            # 命名依赖: "q:0,k:1,v:2"
            for part in parse_list(depends_str):
                if ":" not in part:
                    continue
                name, idx_str = part.split(":", 1)
                name = name.strip()
                try:
                    idx = int(idx_str.strip())
                    result[name] = [idx]
                except ValueError:
                    logger.warning(f"无效的依赖索引: {part}")
        else:
            # 简单依赖: "0" 或 "1,2"
            parts = parse_list(depends_str)
            if len(parts) == 1:
                # 单依赖
                try:
                    result["x"] = [int(parts[0])]
                except ValueError:
                    logger.warning(f"无效的依赖索引: {parts[0]}")
            else:
                # 双输入依赖
                input_names = ["a", "b", "c", "d", "e", "f"]
                for i, part in enumerate(parts):
                    try:
                        name = input_names[i] if i < len(input_names) else f"in{i}"
                        result[name] = [int(part)]
                    except ValueError:
                        logger.warning(f"无效的依赖索引: {part}")

        return result


def _get_str(row_dict: Dict, key: str, default: str = "") -> str:
    """安全获取字符串值"""
    return str(row_dict.get(key, default) or default)


def _parse_op_registry(wb) -> List[str]:
    """解析 op_registry sheet，返回启用的算子列表"""
    if "op_registry" not in wb.sheetnames:
        return []

    ws = wb["op_registry"]
    headers = [cell.value for cell in ws[1]]
    enabled_ops = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        row_dict = dict(zip(headers, row))
        op_name = row_dict.get("op_name", "")
        enabled = str(row_dict.get("enabled", "TRUE")).upper()
        if op_name and enabled == "TRUE":
            enabled_ops.append(op_name)

    return enabled_ops


def _row_to_opconfig(row_dict: Dict) -> OpConfig:
    """将行数据转换为 OpConfig"""
    shape_str = _get_str(row_dict, "shape")
    shape = parse_shape(shape_str) or ()

    skip_str = _get_str(row_dict, "skip", "FALSE").upper()
    skip = skip_str in ("TRUE", "1", "YES")

    return OpConfig(
        id=int(row_dict.get("id", 0) or 0),
        op_name=_get_str(row_dict, "op_name"),
        shape=shape,
        dtype=_get_str(row_dict, "dtype", "float32"),
        depends=_get_str(row_dict, "depends"),
        qtype=_get_str(row_dict, "qtype"),
        skip=skip,
        note=_get_str(row_dict, "note"),
        paths=BinaryPaths(
            golden=_get_str(row_dict, "golden_bin"),
            result=_get_str(row_dict, "result_bin"),
            input=_get_str(row_dict, "input_bin"),
            weight=_get_str(row_dict, "weight_bin"),
            sim_cmd=_get_str(row_dict, "sim_cmd"),
        ),
    )


def _parse_ops_sheet(wb) -> List[OpConfig]:
    """解析 ops sheet，返回算子配置列表"""
    if "ops" not in wb.sheetnames:
        return []

    ws = wb["ops"]
    headers = [cell.value for cell in ws[1]]
    op_configs = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        first_cell = str(row[0]).strip()
        if first_cell.startswith("#"):
            continue

        row_dict = dict(zip(headers, row))
        op_configs.append(_row_to_opconfig(row_dict))

    return op_configs


def parse_xlsx(xlsx_path: str) -> Tuple[List[str], List[OpConfig]]:
    """
    解析 xlsx 文件

    Args:
        xlsx_path: xlsx 文件路径

    Returns:
        (enabled_ops, op_configs)
        - enabled_ops: 启用的算子列表
        - op_configs: 算子配置列表
    """
    _check_openpyxl()

    wb = load_workbook(xlsx_path, data_only=True)
    enabled_ops = _parse_op_registry(wb)
    op_configs = _parse_ops_sheet(wb)
    wb.close()

    return enabled_ops, op_configs


def import_xlsx(xlsx_path: str, output_py: Optional[str] = None) -> str:
    """
    从 xlsx 生成 Python 代码

    Args:
        xlsx_path: xlsx 文件路径
        output_py: 输出 Python 文件路径，None 表示只返回代码不写文件

    Returns:
        生成的 Python 代码
    """
    _, op_configs = parse_xlsx(xlsx_path)

    # 生成代码
    lines = [
        '"""自动生成的算子测试代码',
        "",
        f"从 xlsx 配置生成: {Path(xlsx_path).name}",
        '"""',
        "import numpy as np",
        "from aidevtools.ops import _functional as F",
        "from aidevtools.ops.base import clear, dump, gen_csv",
        "",
        "",
        "def run():",
        '    """执行算子测试"""',
        "    clear()  # 清空之前的记录",
        "    outputs = {}  # 保存各步骤的输出",
        "",
    ]

    for config in op_configs:
        if config.skip:
            lines.append(f"    # [SKIP] {config.op_name} (id={config.id})")
            continue

        indent = "    "
        op_name = config.op_name
        shape = config.shape
        dtype = config.dtype

        # 解析依赖
        depends = config.parse_depends()

        # 生成注释
        if config.note:
            lines.append(f"{indent}# {config.note}")

        # 生成输入
        if not depends:
            # 无依赖，随机输入
            shape_str = ", ".join(str(d) for d in shape) if shape else "1, 64"
            lines.append(f"{indent}x_{config.id} = np.random.randn({shape_str}).astype(np.{dtype})")
            input_var = f"x_{config.id}"
        else:
            # 有依赖
            input_vars = []
            for name, deps in depends.items():
                for dep_id in deps:
                    lines.append(f"{indent}{name}_{config.id} = outputs[{dep_id}]")
                    input_vars.append(f"{name}_{config.id}")
            input_var = (
                ", ".join(input_vars)
                if len(input_vars) > 1
                else input_vars[0]
                if input_vars
                else f"x_{config.id}"
            )

            # 生成算子调用
        _generate_op_call(lines, indent, config, op_name, shape, dtype, depends, input_var)

        # 保存输出
        lines.append(f"{indent}outputs[{config.id}] = out_{config.id}")
        lines.append("")

    # 生成导出代码
    lines.extend(
        [
            "    # 导出结果",
            '    dump("./workspace")',
            '    csv_path = gen_csv("./workspace", "generated")',
            '    print(f"生成 compare 配置: {csv_path}")',
            "    return outputs",
            "",
            "",
            'if __name__ == "__main__":',
            "    run()",
            "",
        ]
    )

    code = "\n".join(lines)

    # 写文件
    if output_py:
        output_path = Path(output_py)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(code, encoding="utf-8")
        logger.info(f"生成 Python 代码: {output_py}")

    return code
