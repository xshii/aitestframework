#!/usr/bin/env python
"""xlsx → Python Transformer Demo

演示从 Excel 配置生成 Transformer 模型并运行:
1. 创建 xlsx 模板
2. 在 xlsx 中定义 Transformer 算子序列
3. 从 xlsx 生成 Python 代码
4. 运行并比对结果

使用方法:
    cd demos/05_xlsx_transformer
    python run.py
"""
from pathlib import Path

from aidevtools.ops.base import set_golden_mode

# 设置 cpu golden dtype 并使用 cpp golden
from aidevtools.ops.cpu_golden import set_cpu_golden_dtype
set_cpu_golden_dtype("gfp16")
set_golden_mode("cpp")


def create_transformer_xlsx(xlsx_path: str):
    """
    创建 Transformer 模型的 xlsx 配置

    模型结构 (简化版 1 层 Transformer):
        input → matmul(Q) → matmul(K) → matmul(V)
              → attention → matmul(O) → add(residual)
              → layernorm → matmul(FFN_up) → softmax
              → matmul(FFN_down) → add(residual) → layernorm

    注意: 使用支持 cpp golden 的算子 (matmul, softmax, layernorm)
    """
    from aidevtools.xlsx import create_template

    # 创建模板，指定支持 cpp golden 的算子
    create_template(xlsx_path, ops=[
        "matmul", "softmax", "layernorm"
    ])

    # 编辑 xlsx 添加 transformer 算子配置
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws_ops = wb["ops"]

    # 清空示例数据
    ws_ops.delete_rows(3, ws_ops.max_row)

    # Transformer 配置 (简化版，使用支持 cpp golden 的算子)
    # 注意：xlsx run 目前仅支持线性依赖链，复杂的多输入操作（如 attention）需要用 Python API
    # 格式: id, op_name, shape, dtype, depends, qtype, skip, sim_cmd, note
    transformer_ops = [
        # 输入层
        (0, "matmul", "1,16,64", "float32", "", "bfp8", "FALSE", "", "Input Projection"),

        # Self-Attention (简化: 只展示单路径)
        (1, "matmul", "1,16,64", "float32", "0", "bfp8", "FALSE", "", "Q Projection"),
        (2, "softmax", "1,16,64", "float32", "1", "bfp8", "FALSE", "", "Attention Weights"),
        (3, "matmul", "1,16,64", "float32", "2", "bfp8", "FALSE", "", "Attention Output"),

        # LayerNorm 1
        (4, "layernorm", "1,16,64", "float32", "3", "bfp8", "FALSE", "", "LayerNorm 1"),

        # FFN
        (5, "matmul", "1,16,256", "float32", "4", "bfp8", "FALSE", "", "FFN Up"),
        (6, "softmax", "1,16,256", "float32", "5", "bfp8", "FALSE", "", "FFN Activation"),
        (7, "matmul", "1,16,64", "float32", "6", "bfp8", "FALSE", "", "FFN Down"),

        # LayerNorm 2
        (8, "layernorm", "1,16,64", "float32", "7", "bfp8", "FALSE", "", "Output LayerNorm"),
    ]

    for row_idx, config in enumerate(transformer_ops, 3):
        for col_idx, value in enumerate(config, 1):
            ws_ops.cell(row=row_idx, column=col_idx, value=value)

    wb.save(xlsx_path)
    print(f"    Transformer 配置已写入: {xlsx_path}")
    print(f"    共 {len(transformer_ops)} 个算子")
    print("    使用 cpp golden (gfp16) + bfp8 量化")

    return transformer_ops


def show_xlsx_content(xlsx_path: str):
    """显示 xlsx 内容"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)

    print("\n    [ops sheet - Transformer 算子配置]")
    print("    " + "-" * 80)

    ws_ops = wb["ops"]
    # 打印表头
    headers = [cell.value for cell in ws_ops[1]]
    print(f"    {headers}")
    print("    " + "-" * 80)

    # 打印数据
    for row in ws_ops.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:  # 跳过空行
            # 格式化输出
            row_str = f"    [{row[0]:2}] {row[1]:12} shape={row[2]:12} qtype={row[5]:6} note={row[8]}"
            print(row_str)


def generate_python_code(xlsx_path: str, output_dir: str) -> str:
    """从 xlsx 生成 Python 代码"""
    from aidevtools.xlsx import import_xlsx

    py_path = Path(output_dir) / "generated_transformer.py"
    code = import_xlsx(xlsx_path, str(py_path))

    print(f"\n    生成 Python 代码: {py_path}")
    print("\n    [生成的代码预览]")
    print("    " + "-" * 60)

    lines = code.split("\n")
    for i, line in enumerate(lines[:30]):
        print(f"    {line}")
    if len(lines) > 30:
        print(f"    ... (共 {len(lines)} 行)")

    return str(py_path)


def run_transformer_xlsx(xlsx_path: str, output_dir: str):
    """运行 xlsx 配置的 Transformer"""
    from aidevtools.xlsx import run_xlsx

    print("\n    运行 Transformer 模型 (cpp golden)...")
    results = run_xlsx(xlsx_path, output_dir)

    print("\n    [运行结果]")
    print("    " + "-" * 60)
    print(f"    {'ID':>3} {'算子':12} {'状态':8} {'备注'}")
    print("    " + "-" * 60)

    pass_count = 0
    fail_count = 0
    skip_count = 0

    for r in results:
        status = r.get("status", "?")
        op_id = r.get("id", "?")
        op_name = r.get("op_name", "?")
        note = r.get("note", "")

        if status == "PASS":
            status_icon = "✓"
            pass_count += 1
        elif status == "FAIL":
            status_icon = "✗"
            fail_count += 1
        elif status == "SKIP":
            status_icon = "○"
            skip_count += 1
        else:
            status_icon = "?"

        print(f"    [{status_icon}] {op_id:2} {op_name:12} {status:8} {note}")

    print("    " + "-" * 60)
    print(f"    总计: {pass_count} PASS, {fail_count} FAIL, {skip_count} SKIP")

    return results


def main():
    """主函数"""
    print("""
╔══════════════════════════════════════════════════════════════════════╗
║              xlsx → Python Transformer Demo                          ║
╠══════════════════════════════════════════════════════════════════════╣
║  从 Excel 配置生成 Transformer 模型:                                 ║
║  1. xlsx 定义算子序列和量化类型                                      ║
║  2. 自动生成 Python 代码                                             ║
║  3. 运行模型并比对结果                                               ║
║                                                                      ║
║  使用 cpp golden (via subprocess)                                    ║
║  量化: gfp16 (cpp) + bfp8 (input)                                    ║
╚══════════════════════════════════════════════════════════════════════╝
""")

    # 检查 openpyxl
    try:
        import openpyxl
    except ImportError:
        print("错误: 需要安装 openpyxl")
        print("  pip install openpyxl")
        return

    output_dir = Path(__file__).parent / "workspace"
    output_dir.mkdir(exist_ok=True)
    xlsx_path = output_dir / "transformer_config.xlsx"

    # Step 1: 创建 Transformer xlsx 配置
    print("[Step 1] 创建 Transformer xlsx 配置")
    print("-" * 50)
    create_transformer_xlsx(str(xlsx_path))

    # Step 2: 显示 xlsx 内容
    print("\n[Step 2] xlsx 内容")
    print("-" * 50)
    show_xlsx_content(str(xlsx_path))

    # Step 3: 生成 Python 代码
    print("\n[Step 3] 生成 Python 代码")
    print("-" * 50)
    py_path = generate_python_code(str(xlsx_path), str(output_dir))

    # Step 4: 运行 Transformer
    print("\n[Step 4] 运行 Transformer 并比对")
    print("-" * 50)
    run_transformer_xlsx(str(xlsx_path), str(output_dir))

    # 总结
    print("\n" + "=" * 70)
    print("Demo 完成!")
    print("=" * 70)
    print(f"""
文件位置:
  xlsx 配置:    {xlsx_path}
  Python 代码:  {py_path}
  输出目录:     {output_dir}

Golden 配置:
  - cpp golden via subprocess (gfp16)
  - 输入量化: bfp8

模型结构 (简化版，线性依赖):
  matmul (Input) → matmul (Q) → softmax → matmul (Attn)
        → layernorm → FFN matmul → softmax → FFN matmul
        → layernorm (Output)
""")


if __name__ == "__main__":
    main()
