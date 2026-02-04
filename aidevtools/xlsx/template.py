"""xlsx 模板生成

生成包含三个 sheet 的空模板:
1. op_registry - 可用算子列表
2. ops - 用户配置算子用例
3. compare - 比数结果 (自动生成)
"""

from pathlib import Path
from typing import List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from aidevtools.core.log import logger
from aidevtools.xlsx.op_registry import get_default_ops, get_op_info


def _check_openpyxl():
    """检查 openpyxl 是否可用"""
    if not HAS_OPENPYXL:
        raise ImportError("xlsx 功能需要 openpyxl，请安装: pip install openpyxl")


def create_template(
    output_path: str,
    ops: Optional[List[str]] = None,
    include_examples: bool = True,
) -> str:
    """
    创建 xlsx 模板

    Args:
        output_path: 输出文件路径
        ops: 限定的算子列表，None 表示使用全部默认算子
        include_examples: 是否包含示例行

    Returns:
        输出文件路径
    """
    _check_openpyxl()

    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ==================== Sheet 1: op_registry ====================
    ws_registry = wb.active
    ws_registry.title = "op_registry"

    # 表头
    registry_headers = ["op_name", "enabled", "inputs", "optional", "description"]
    for col, header in enumerate(registry_headers, 1):
        cell = ws_registry.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # 数据
    default_ops = get_default_ops()
    use_ops = ops if ops else list(default_ops.keys())

    for row, op_name in enumerate(use_ops, 2):
        info = get_op_info(op_name)
        ws_registry.cell(row=row, column=1, value=op_name).border = thin_border
        ws_registry.cell(row=row, column=2, value="TRUE").border = thin_border
        ws_registry.cell(
            row=row, column=3, value=",".join(info.get("inputs", []))
        ).border = thin_border
        ws_registry.cell(
            row=row, column=4, value=",".join(info.get("optional", []))
        ).border = thin_border
        ws_registry.cell(row=row, column=5, value=info.get("description", "")).border = thin_border

    # 列宽
    ws_registry.column_dimensions["A"].width = 15
    ws_registry.column_dimensions["B"].width = 10
    ws_registry.column_dimensions["C"].width = 20
    ws_registry.column_dimensions["D"].width = 25
    ws_registry.column_dimensions["E"].width = 40

    # ==================== Sheet 2: ops ====================
    ws_ops = wb.create_sheet("ops")

    # 表头
    ops_headers = [
        "id",
        "op_name",
        "shape",
        "dtype",
        "depends",
        "qtype",
        "skip",
        "sim_cmd",
        "golden_bin",
        "result_bin",
        "input_bin",
        "weight_bin",  # binary 路径
        "note",
    ]
    for col, header in enumerate(ops_headers, 1):
        cell = ws_ops.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # 示例数据（包含说明行）
    if include_examples:
        # 添加说明行
        help_text = (
            "# depends: 空=随机,0=依赖第0行,\"1,2\"=双输入,\"q:0,k:1,v:2\"=命名; "
            "*_bin: 留空=自动生成,填写=使用指定路径; "
            "sim_cmd占位符: {golden_bin},{result_bin},{input_bin},{weight_bin},{id},{op_name}"
        )
        ws_ops.cell(row=2, column=1, value=help_text)
        ws_ops.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
        ws_ops.cell(row=2, column=1).font = Font(italic=True, color="666666")
        examples = [
            {
                "id": 0,
                "op_name": "linear",
                "shape": "1,64,128",
                "dtype": "float32",
                "depends": "",
                "qtype": "",
                "skip": "FALSE",
                "sim_cmd": "",
                "golden_bin": "",
                "result_bin": "",
                "input_bin": "",
                "weight_bin": "",
                "note": "binary路径留空=自动生成",
            },
            {
                "id": 1,
                "op_name": "relu",
                "shape": "1,64,256",
                "dtype": "float32",
                "depends": "0",
                "qtype": "",
                "skip": "FALSE",
                "sim_cmd": "",
                "golden_bin": "",
                "result_bin": "",
                "input_bin": "",
                "weight_bin": "",
                "note": "依赖第0行输出",
            },
            {
                "id": 2,
                "op_name": "attention",
                "shape": "1,8,64,64",
                "dtype": "float32",
                "depends": "q:0,k:0,v:1",
                "qtype": "",
                "skip": "FALSE",
                "sim_cmd": "",
                "golden_bin": "",
                "result_bin": "",
                "input_bin": "",
                "weight_bin": "",
                "note": "命名依赖",
            },
            {
                "id": 3,
                "op_name": "matmul",
                "shape": "1,64,64",
                "dtype": "float32",
                "depends": "1,2",
                "qtype": "float16",
                "skip": "FALSE",
                "sim_cmd": "./sim.sh {input_bin} {golden_bin} {result_bin}",
                "golden_bin": "",
                "result_bin": "/path/to/dut_result.bin",
                "input_bin": "",
                "weight_bin": "",
                "note": "指定DUT输出路径",
            },
        ]
        for row_idx, ex in enumerate(examples, 3):
            ws_ops.cell(row=row_idx, column=1, value=ex["id"]).border = thin_border
            ws_ops.cell(row=row_idx, column=2, value=ex["op_name"]).border = thin_border
            ws_ops.cell(row=row_idx, column=3, value=ex["shape"]).border = thin_border
            ws_ops.cell(row=row_idx, column=4, value=ex["dtype"]).border = thin_border
            ws_ops.cell(row=row_idx, column=5, value=ex["depends"]).border = thin_border
            ws_ops.cell(row=row_idx, column=6, value=ex["qtype"]).border = thin_border
            ws_ops.cell(row=row_idx, column=7, value=ex["skip"]).border = thin_border
            ws_ops.cell(row=row_idx, column=8, value=ex["sim_cmd"]).border = thin_border
            ws_ops.cell(row=row_idx, column=9, value=ex["golden_bin"]).border = thin_border
            ws_ops.cell(row=row_idx, column=10, value=ex["result_bin"]).border = thin_border
            ws_ops.cell(row=row_idx, column=11, value=ex["input_bin"]).border = thin_border
            ws_ops.cell(row=row_idx, column=12, value=ex["weight_bin"]).border = thin_border
            ws_ops.cell(row=row_idx, column=13, value=ex["note"]).border = thin_border

    # 列宽
    ws_ops.column_dimensions["A"].width = 8  # id
    ws_ops.column_dimensions["B"].width = 15  # op_name
    ws_ops.column_dimensions["C"].width = 20  # shape
    ws_ops.column_dimensions["D"].width = 12  # dtype
    ws_ops.column_dimensions["E"].width = 20  # depends
    ws_ops.column_dimensions["F"].width = 12  # qtype
    ws_ops.column_dimensions["G"].width = 8  # skip
    ws_ops.column_dimensions["H"].width = 40  # sim_cmd
    ws_ops.column_dimensions["I"].width = 35  # golden_bin
    ws_ops.column_dimensions["J"].width = 35  # result_bin
    ws_ops.column_dimensions["K"].width = 35  # input_bin
    ws_ops.column_dimensions["L"].width = 35  # weight_bin
    ws_ops.column_dimensions["M"].width = 30  # note

    # ==================== Sheet 3: compare ====================
    ws_compare = wb.create_sheet("compare")

    # 表头
    compare_headers = [
        "id",
        "op_name",
        "status",
        "max_abs",
        "qsnr",
        "cosine",
        "isclose_pass",
        "exceed_count",
        "exceed_ratio",
        "golden_bin",
        "result_bin",
        "note",
    ]
    for col, header in enumerate(compare_headers, 1):
        cell = ws_compare.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # 列宽
    ws_compare.column_dimensions["A"].width = 8
    ws_compare.column_dimensions["B"].width = 15
    ws_compare.column_dimensions["C"].width = 10
    ws_compare.column_dimensions["D"].width = 15
    ws_compare.column_dimensions["E"].width = 12
    ws_compare.column_dimensions["F"].width = 12
    ws_compare.column_dimensions["G"].width = 40
    ws_compare.column_dimensions["H"].width = 40
    ws_compare.column_dimensions["I"].width = 30

    # 保存
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    logger.info(f"创建 xlsx 模板: {output_path}")
    return str(output_path)
