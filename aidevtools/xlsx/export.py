"""xlsx 导出

从 trace 记录导出到 xlsx，保留已有的结果列。
"""

from pathlib import Path
from typing import Any, Dict, List

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Font, PatternFill, Side

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from aidevtools.core.log import logger


def _check_openpyxl():
    if not HAS_OPENPYXL:
        raise ImportError("xlsx 功能需要 openpyxl，请安装: pip install openpyxl")


def _load_existing_compare_data(output_path: Path) -> Dict[Any, Dict]:
    """加载已有的 compare 结果"""
    existing = {}
    try:
        wb = load_workbook(output_path)
        if "compare" in wb.sheetnames:
            ws = wb["compare"]
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if row_dict.get("id") is not None:
                    existing[row_dict["id"]] = row_dict
            logger.debug(f"保留 {len(existing)} 条已有比对结果")
        wb.close()
    except (OSError, KeyError, ValueError, TypeError) as e:
        logger.warning(f"读取已有结果失败: {e}")
    return existing


def _get_xlsx_styles():
    """获取 xlsx 样式"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    return header_font, header_fill, thin_border


def _write_ops_sheet(wb, records: List[Dict], header_font, header_fill, border):
    """写入 ops sheet"""
    import numpy as np

    ops_headers = ["id", "op_name", "shape", "dtype", "depends", "qtype", "skip", "note"]

    if "ops" in wb.sheetnames:
        ws = wb["ops"]
        ws.delete_rows(2, ws.max_row)
    else:
        ws = wb.create_sheet("ops")
        for col, header in enumerate(ops_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font, cell.fill, cell.border = header_font, header_fill, border

    for idx, record in enumerate(records):
        row = idx + 2
        name = record.get("name", f"op_{idx}")
        op = record.get("op", name.rsplit("_", 1)[0])

        output = record.get("golden")
        if output is None:
            output = record.get("output")
        shape_str, dtype_str = "", ""
        if output is not None and hasattr(output, "shape"):
            arr = np.asarray(output)
            shape_str = ",".join(str(d) for d in arr.shape)
            dtype_str = str(arr.dtype)

        values = [idx, op, shape_str, dtype_str, "", "", "FALSE", name]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val).border = border


def _write_compare_sheet(
    wb, records: List[Dict], existing_data: Dict, header_font, header_fill, border
):
    """写入 compare sheet"""
    compare_headers = [
        "id",
        "op_name",
        "status",
        "max_abs",
        "qsnr",
        "cosine",
        "golden_bin",
        "result_bin",
        "note",
    ]
    compare_keys = ["status", "max_abs", "qsnr", "cosine", "golden_bin", "result_bin", "note"]

    if "compare" in wb.sheetnames:
        ws = wb["compare"]
        ws.delete_rows(2, ws.max_row)
    else:
        ws = wb.create_sheet("compare")
        for col, header in enumerate(compare_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font, cell.fill, cell.border = header_font, header_fill, border

    for idx, record in enumerate(records):
        row = idx + 2
        name = record.get("name", f"op_{idx}")
        existing = existing_data.get(idx, {})

        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=name).border = border
        for col, key in enumerate(compare_keys, 3):
            ws.cell(row=row, column=col, value=existing.get(key, "")).border = border


def export_xlsx(
    output_path: str,
    records: List[Dict[str, Any]],
    preserve_results: bool = True,
) -> str:
    """导出 trace 记录到 xlsx"""
    _check_openpyxl()

    output_path = Path(output_path)
    existing_data = (
        _load_existing_compare_data(output_path)
        if preserve_results and output_path.exists()
        else {}
    )

    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        from aidevtools.xlsx.template import create_template

        create_template(str(output_path), include_examples=False)
        wb = load_workbook(output_path)

    header_font, header_fill, border = _get_xlsx_styles()
    _write_ops_sheet(wb, records, header_font, header_fill, border)
    _write_compare_sheet(wb, records, existing_data, header_font, header_fill, border)

    wb.save(output_path)
    logger.info(f"导出 xlsx: {output_path} ({len(records)} 条记录)")
    return str(output_path)


def _build_id_to_row_map(ws, col_map: Dict[str, int]) -> Dict[Any, int]:
    """构建 id -> row 映射"""
    id_to_row = {}
    id_col = col_map.get("id", 1)
    for row in range(2, ws.max_row + 1):
        cell_id = ws.cell(row=row, column=id_col).value
        if cell_id is not None:
            id_to_row[cell_id] = row
    return id_to_row


def _update_cell_with_style(ws, row: int, col: int, value: str, pass_fill, fail_fill):
    """更新单元格并应用状态样式"""
    cell = ws.cell(row=row, column=col, value=value)
    if value == "PASS":
        cell.fill = pass_fill
    elif value == "FAIL":
        cell.fill = fail_fill


def update_compare_results(
    xlsx_path: str,
    results: List[Dict[str, Any]],
) -> str:
    """
    更新 xlsx 中的比对结果

    Args:
        xlsx_path: xlsx 文件路径
        results: 比对结果列表，每项包含 id, status, max_abs, qsnr, cosine 等

    Returns:
        xlsx 文件路径
    """
    _check_openpyxl()

    wb = load_workbook(xlsx_path)
    if "compare" not in wb.sheetnames:
        raise ValueError(f"xlsx 文件缺少 compare sheet: {xlsx_path}")

    ws = wb["compare"]
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers)}
    id_to_row = _build_id_to_row_map(ws, col_map)

    # 样式
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 普通字段列表 (直接更新值)
    value_fields = [
        "max_abs",
        "qsnr",
        "cosine",
        "golden_bin",
        "result_bin",
        "note",
        "exceed_count",
        "exceed_ratio",
    ]
    # 状态字段列表 (需要应用样式)
    status_fields = ["status", "isclose_pass"]

    for res in results:
        res_id = res.get("id")
        if res_id not in id_to_row:
            continue

        row = id_to_row[res_id]

        # 更新状态字段 (带样式)
        for field in status_fields:
            if field in col_map and field in res:
                _update_cell_with_style(ws, row, col_map[field], res[field], pass_fill, fail_fill)

        # 更新普通字段
        for field in value_fields:
            if field in col_map and field in res:
                ws.cell(row=row, column=col_map[field], value=res[field])

    wb.save(xlsx_path)
    logger.info(f"更新比对结果: {xlsx_path}")
    return xlsx_path
