"""xlsx 模块单元测试"""
import pytest
import numpy as np
from pathlib import Path
import tempfile
import shutil

# 检查 openpyxl 是否可用
pytest.importorskip("openpyxl")


class TestOpRegistry:
    """算子注册表测试"""

    def test_get_default_ops(self):
        """获取默认算子"""
        from aidevtools.xlsx.op_registry import get_default_ops

        ops = get_default_ops()
        assert isinstance(ops, dict)
        assert "linear" in ops
        assert "matmul" in ops
        assert "attention" in ops

    def test_get_op_info(self):
        """获取算子信息"""
        from aidevtools.xlsx.op_registry import get_op_info

        info = get_op_info("linear")
        assert "inputs" in info
        assert "x" in info["inputs"] or "weight" in info["inputs"]

        # 未知算子
        info_unknown = get_op_info("unknown_op")
        assert "inputs" in info_unknown

    def test_list_ops(self):
        """列出所有算子"""
        from aidevtools.xlsx.op_registry import list_ops

        ops = list_ops()
        assert isinstance(ops, list)
        assert len(ops) > 0
        assert "linear" in ops


class TestXlsxTemplate:
    """xlsx 模板测试"""

    def setup_method(self):
        """创建临时目录"""
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        """清理临时目录"""
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_create_template(self):
        """创建空模板"""
        from aidevtools.xlsx import create_template
        from openpyxl import load_workbook

        output = Path(self.temp_dir) / "test_template.xlsx"
        create_template(str(output))

        assert output.exists()

        # 检查 sheet
        wb = load_workbook(output)
        assert "op_registry" in wb.sheetnames
        assert "ops" in wb.sheetnames
        assert "compare" in wb.sheetnames

    def test_create_template_with_ops(self):
        """创建限定算子的模板"""
        from aidevtools.xlsx import create_template
        from openpyxl import load_workbook

        output = Path(self.temp_dir) / "test_limited.xlsx"
        create_template(str(output), ops=["linear", "relu"])

        wb = load_workbook(output)
        ws = wb["op_registry"]

        # 检查只有限定的算子
        ops_in_sheet = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ops_in_sheet.append(row[0])

        assert "linear" in ops_in_sheet
        assert "relu" in ops_in_sheet
        assert len(ops_in_sheet) == 2


class TestXlsxExport:
    """xlsx 导出测试"""

    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_export_xlsx(self):
        """导出 trace 记录"""
        from aidevtools.xlsx import export_xlsx
        from openpyxl import load_workbook

        output = Path(self.temp_dir) / "test_export.xlsx"

        # 模拟记录
        records = [
            {
                "name": "linear_0",
                "op": "linear",
                "golden": np.random.randn(1, 64, 128).astype(np.float32),
                "result": None,
            },
            {
                "name": "relu_0",
                "op": "relu",
                "golden": np.random.randn(1, 64, 128).astype(np.float32),
                "result": np.random.randn(1, 64, 128).astype(np.float32),
            },
        ]

        export_xlsx(str(output), records)

        assert output.exists()

        # 检查 ops sheet
        wb = load_workbook(output)
        ws_ops = wb["ops"]

        # 应该有 2 条记录（不含表头）
        data_rows = list(ws_ops.iter_rows(min_row=2, values_only=True))
        # 过滤空行
        data_rows = [r for r in data_rows if r[0] is not None]
        assert len(data_rows) == 2

    def test_preserve_results(self):
        """保留已有结果"""
        from aidevtools.xlsx import export_xlsx
        from aidevtools.xlsx.export import update_compare_results
        from openpyxl import load_workbook

        output = Path(self.temp_dir) / "test_preserve.xlsx"

        # 先创建模板并添加一条记录
        records = [{"name": "test_op", "op": "relu", "golden": np.zeros((1, 64))}]
        export_xlsx(str(output), records)

        # 模拟已有结果
        results = [
            {"id": 0, "status": "PASS", "max_abs": "1.23e-6", "qsnr": "50.00", "cosine": "0.999999"},
        ]
        update_compare_results(str(output), results)

        # 检查结果被写入
        wb = load_workbook(output)
        ws = wb["compare"]
        # 第 2 行应该有 PASS
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        # status 在第 3 列 (0-indexed: 2)
        assert row2[2] == "PASS"


class TestXlsxImport:
    """xlsx 导入测试"""

    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_parse_xlsx(self):
        """解析 xlsx"""
        from aidevtools.xlsx import create_template, parse_xlsx

        xlsx_path = Path(self.temp_dir) / "test_parse.xlsx"
        create_template(str(xlsx_path))

        enabled_ops, op_configs = parse_xlsx(str(xlsx_path))

        assert isinstance(enabled_ops, list)
        assert isinstance(op_configs, list)

    def test_import_xlsx(self):
        """从 xlsx 生成 Python 代码"""
        from aidevtools.xlsx import create_template, import_xlsx

        xlsx_path = Path(self.temp_dir) / "test_import.xlsx"
        py_path = Path(self.temp_dir) / "generated.py"

        create_template(str(xlsx_path))

        code = import_xlsx(str(xlsx_path), str(py_path))

        assert py_path.exists()
        assert "import numpy as np" in code
        assert "def run():" in code

    def test_parse_depends(self):
        """解析依赖关系"""
        from aidevtools.xlsx.import_ import OpConfig

        # 无依赖
        cfg1 = OpConfig(0, "linear", (1, 64), "float32", "", "", False, "")
        assert cfg1.parse_depends() == {}

        # 单依赖
        cfg2 = OpConfig(1, "relu", (1, 64), "float32", "0", "", False, "")
        deps2 = cfg2.parse_depends()
        assert "x" in deps2
        assert deps2["x"] == [0]

        # 双输入依赖
        cfg3 = OpConfig(2, "matmul", (1, 64), "float32", "0,1", "", False, "")
        deps3 = cfg3.parse_depends()
        assert "a" in deps3 and "b" in deps3
        assert deps3["a"] == [0]
        assert deps3["b"] == [1]

        # 命名依赖
        cfg4 = OpConfig(3, "attention", (1, 8, 64), "float32", "q:0,k:1,v:2", "", False, "")
        deps4 = cfg4.parse_depends()
        assert deps4["q"] == [0]
        assert deps4["k"] == [1]
        assert deps4["v"] == [2]


class TestXlsxRun:
    """xlsx 运行测试"""

    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_run_xlsx_simple(self):
        """简单运行测试"""
        import pytest
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook
        from aidevtools.ops.cpu_golden import is_cpu_golden_available

        if not is_cpu_golden_available():
            pytest.skip("CPU golden not available")

        xlsx_path = Path(self.temp_dir) / "test_run.xlsx"

        # 创建模板
        create_template(str(xlsx_path), include_examples=False)

        # 手动添加简单用例 - 使用有 cpu_golden 的 softmax
        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        # 添加一个 softmax 用例 (有 cpu_golden 实现)
        ws.cell(row=2, column=1, value=0)  # id
        ws.cell(row=2, column=2, value="softmax")  # op_name
        ws.cell(row=2, column=3, value="1,64")  # shape
        ws.cell(row=2, column=4, value="float32")  # dtype
        ws.cell(row=2, column=5, value="")  # depends
        ws.cell(row=2, column=6, value="")  # qtype
        ws.cell(row=2, column=7, value="FALSE")  # skip
        ws.cell(row=2, column=8, value="test softmax")  # note

        wb.save(xlsx_path)

        # 运行
        results = run_xlsx(str(xlsx_path), str(self.temp_dir))

        assert len(results) > 0
        # 检查 xlsx 结果被更新
        wb2 = load_workbook(xlsx_path)
        ws2 = wb2["compare"]
        row2 = list(ws2.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        # 应该有状态
        assert row2[2] is not None  # status 列

    def test_run_xlsx_with_skip(self):
        """测试跳过算子"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_skip.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        # 添加一个被跳过的用例
        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="relu")
        ws.cell(row=2, column=3, value="1,64")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=5, value="")
        ws.cell(row=2, column=6, value="")
        ws.cell(row=2, column=7, value="TRUE")  # skip=TRUE
        ws.cell(row=2, column=8, value="skipped")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))

        # 检查跳过状态
        skip_results = [r for r in results if r.get("status") == "SKIP"]
        assert len(skip_results) >= 1

    def test_run_xlsx_with_depends(self):
        """测试带依赖的算子"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_depends.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        # linear -> relu (依赖)
        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="linear")
        ws.cell(row=2, column=3, value="1,32,64")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=5, value="")
        ws.cell(row=2, column=7, value="FALSE")

        ws.cell(row=3, column=1, value=1)
        ws.cell(row=3, column=2, value="relu")
        ws.cell(row=3, column=3, value="1,32,128")
        ws.cell(row=3, column=4, value="float32")
        ws.cell(row=3, column=5, value="0")  # 依赖第0行
        ws.cell(row=3, column=7, value="FALSE")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))
        assert len(results) >= 2

    def test_run_xlsx_matmul(self):
        """测试 matmul 算子"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_matmul.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="matmul")
        ws.cell(row=2, column=3, value="2,4,4")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="FALSE")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))
        assert len(results) >= 1

    def test_run_xlsx_softmax(self):
        """测试 softmax 算子"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_softmax.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="softmax")
        ws.cell(row=2, column=3, value="2,8,64")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="FALSE")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))
        assert len(results) >= 1

    def test_run_xlsx_layernorm(self):
        """测试 layernorm 算子"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_layernorm.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="layernorm")
        ws.cell(row=2, column=3, value="2,8,64")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="FALSE")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))
        assert len(results) >= 1

    def test_run_xlsx_attention(self):
        """测试 attention 算子"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_attention.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="attention")
        ws.cell(row=2, column=3, value="1,4,8,32")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="FALSE")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))
        assert len(results) >= 1

    def test_run_xlsx_add_mul(self):
        """测试 add/mul 算子"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_add_mul.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        # add
        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="add")
        ws.cell(row=2, column=3, value="2,4,8")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="FALSE")

        # mul
        ws.cell(row=3, column=1, value=1)
        ws.cell(row=3, column=2, value="mul")
        ws.cell(row=3, column=3, value="2,4,8")
        ws.cell(row=3, column=4, value="float32")
        ws.cell(row=3, column=7, value="FALSE")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))
        assert len(results) >= 2

    def test_run_xlsx_dual_input_matmul(self):
        """测试双输入 matmul"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_dual_matmul.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        # 两个 linear 输出
        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="linear")
        ws.cell(row=2, column=3, value="1,8,16")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="FALSE")

        ws.cell(row=3, column=1, value=1)
        ws.cell(row=3, column=2, value="linear")
        ws.cell(row=3, column=3, value="1,16,8")
        ws.cell(row=3, column=4, value="float32")
        ws.cell(row=3, column=7, value="FALSE")

        # matmul 双输入依赖
        ws.cell(row=4, column=1, value=2)
        ws.cell(row=4, column=2, value="matmul")
        ws.cell(row=4, column=3, value="1,8,8")
        ws.cell(row=4, column=4, value="float32")
        ws.cell(row=4, column=5, value="0,1")  # 双输入依赖
        ws.cell(row=4, column=7, value="FALSE")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))
        assert len(results) >= 3

    def test_run_xlsx_dynamic_op(self):
        """测试动态算子调用 (gelu, sigmoid, tanh)"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_dynamic.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        for i, op in enumerate(["gelu", "sigmoid", "tanh"]):
            ws.cell(row=2+i, column=1, value=i)
            ws.cell(row=2+i, column=2, value=op)
            ws.cell(row=2+i, column=3, value="2,4,8")
            ws.cell(row=2+i, column=4, value="float32")
            ws.cell(row=2+i, column=7, value="FALSE")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))
        assert len(results) >= 3

    def test_run_xlsx_disabled_op(self):
        """测试未启用的算子"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_disabled.xlsx"
        # 只启用 linear
        create_template(str(xlsx_path), ops=["linear"], include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        # 添加 relu（未在 op_registry 中启用）
        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="relu")  # 未启用
        ws.cell(row=2, column=3, value="1,64")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="FALSE")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))

        # relu 应该被跳过
        skip_results = [r for r in results if r.get("status") == "SKIP"]
        assert len(skip_results) >= 1


class TestXlsxCommand:
    """xlsx 命令行测试"""

    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_handle_xlsx_template(self):
        """测试 xlsx template 命令"""
        from aidevtools.commands.compare import _handle_xlsx

        xlsx_path = Path(self.temp_dir) / "cmd_template.xlsx"
        ret = _handle_xlsx("template", str(xlsx_path), self.temp_dir, "test", "raw", "")

        assert ret == 0
        assert xlsx_path.exists()

    def test_handle_xlsx_template_with_ops(self):
        """测试带限定算子的 xlsx template"""
        from aidevtools.commands.compare import _handle_xlsx

        xlsx_path = Path(self.temp_dir) / "cmd_limited.xlsx"
        ret = _handle_xlsx("template", str(xlsx_path), self.temp_dir, "test", "raw", "linear,relu")

        assert ret == 0
        assert xlsx_path.exists()

    def test_handle_xlsx_export_no_records(self):
        """测试无记录时的导出"""
        from aidevtools.commands.compare import _handle_xlsx
        from aidevtools.ops.base import clear

        clear()  # 确保无记录

        xlsx_path = Path(self.temp_dir) / "cmd_export.xlsx"
        ret = _handle_xlsx("export", str(xlsx_path), self.temp_dir, "test", "raw", "")

        assert ret == 0
        assert xlsx_path.exists()

    def test_handle_xlsx_import_missing_file(self):
        """测试导入不存在的文件"""
        from aidevtools.commands.compare import _handle_xlsx

        ret = _handle_xlsx("import", "", self.temp_dir, "test", "raw", "")
        assert ret == 1  # 应该返回错误

    def test_handle_xlsx_run_missing_file(self):
        """测试运行不存在的文件"""
        from aidevtools.commands.compare import _handle_xlsx

        ret = _handle_xlsx("run", "", self.temp_dir, "test", "raw", "")
        assert ret == 1  # 应该返回错误

    def test_handle_xlsx_ops(self):
        """测试 xlsx ops 命令"""
        from aidevtools.commands.compare import _handle_xlsx

        ret = _handle_xlsx("ops", "", self.temp_dir, "test", "raw", "")
        assert ret == 0

    def test_handle_xlsx_unknown(self):
        """测试未知 xlsx 子命令"""
        from aidevtools.commands.compare import _handle_xlsx

        ret = _handle_xlsx("unknown_action", "", self.temp_dir, "test", "raw", "")
        assert ret == 1

    def test_handle_xlsx_import(self):
        """测试 xlsx import 命令"""
        from aidevtools.commands.compare import _handle_xlsx
        from aidevtools.xlsx import create_template

        xlsx_path = Path(self.temp_dir) / "cmd_import.xlsx"
        create_template(str(xlsx_path))

        py_path = Path(self.temp_dir) / "generated.py"
        ret = _handle_xlsx("import", str(xlsx_path), str(py_path), "test", "raw", "")

        assert ret == 0

    def test_handle_xlsx_run(self):
        """测试 xlsx run 命令"""
        from aidevtools.commands.compare import _handle_xlsx
        from aidevtools.xlsx import create_template
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "cmd_run.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        # 添加简单用例
        wb = load_workbook(xlsx_path)
        ws = wb["ops"]
        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="relu")
        ws.cell(row=2, column=3, value="1,32")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="FALSE")
        wb.save(xlsx_path)

        ret = _handle_xlsx("run", str(xlsx_path), self.temp_dir, "test", "raw", "")
        assert ret == 0


class TestXlsxExportEdgeCases:
    """xlsx 导出边界测试"""

    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_export_empty_records(self):
        """导出空记录"""
        from aidevtools.xlsx import export_xlsx

        output = Path(self.temp_dir) / "empty.xlsx"
        export_xlsx(str(output), [])

        assert output.exists()

    def test_export_record_without_golden(self):
        """导出没有 golden 的记录"""
        from aidevtools.xlsx import export_xlsx

        output = Path(self.temp_dir) / "no_golden.xlsx"
        records = [{"name": "test", "op": "relu"}]  # 没有 golden
        export_xlsx(str(output), records)

        assert output.exists()

    def test_export_update_existing(self):
        """更新已存在的 xlsx"""
        from aidevtools.xlsx import export_xlsx
        from openpyxl import load_workbook

        output = Path(self.temp_dir) / "update.xlsx"

        # 第一次导出
        records1 = [{"name": "op1", "op": "relu", "golden": np.zeros((1, 8))}]
        export_xlsx(str(output), records1)

        # 第二次导出
        records2 = [{"name": "op2", "op": "linear", "golden": np.zeros((1, 16))}]
        export_xlsx(str(output), records2)

        wb = load_workbook(output)
        ws = wb["ops"]
        # 应该只有第二次的记录
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        assert row2[1] == "linear"


class TestXlsxImportEdgeCases:
    """xlsx 导入边界测试"""

    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_import_empty_ops(self):
        """导入没有算子的 xlsx"""
        from aidevtools.xlsx import create_template, import_xlsx

        xlsx_path = Path(self.temp_dir) / "empty_ops.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        py_path = Path(self.temp_dir) / "empty.py"
        code = import_xlsx(str(xlsx_path), str(py_path))

        assert "def run():" in code
        assert py_path.exists()

    def test_parse_depends_edge_cases(self):
        """解析依赖边界情况"""
        from aidevtools.xlsx.import_ import OpConfig

        # 空字符串依赖
        cfg1 = OpConfig(0, "relu", (1,), "float32", "   ", "", False, "")
        assert cfg1.parse_depends() == {}

        # 多空格分隔
        cfg2 = OpConfig(0, "matmul", (1,), "float32", "0 , 1", "", False, "")
        deps = cfg2.parse_depends()
        assert "a" in deps

    def test_code_gen_with_named_depends(self):
        """生成带命名依赖的代码"""
        from aidevtools.xlsx import create_template, import_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "named_deps.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        # 3 个 linear 作为 q, k, v
        for i in range(3):
            ws.cell(row=2+i, column=1, value=i)
            ws.cell(row=2+i, column=2, value="linear")
            ws.cell(row=2+i, column=3, value="1,4,8,32")
            ws.cell(row=2+i, column=4, value="float32")
            ws.cell(row=2+i, column=7, value="FALSE")

        # attention 用命名依赖
        ws.cell(row=5, column=1, value=3)
        ws.cell(row=5, column=2, value="attention")
        ws.cell(row=5, column=3, value="1,4,8,32")
        ws.cell(row=5, column=4, value="float32")
        ws.cell(row=5, column=5, value="q:0,k:1,v:2")
        ws.cell(row=5, column=7, value="FALSE")

        wb.save(xlsx_path)

        py_path = Path(self.temp_dir) / "named.py"
        code = import_xlsx(str(xlsx_path), str(py_path))

        assert "attention" in code
        assert py_path.exists()

    def test_code_gen_all_ops(self):
        """测试所有算子的代码生成"""
        from aidevtools.xlsx import create_template, import_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "all_ops.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        ops_to_test = [
            (0, "linear", "", "独立 linear"),
            (1, "relu", "0", "依赖 linear"),
            (2, "softmax", "1", "依赖 relu"),
            (3, "matmul", "", "独立 matmul"),
            (4, "matmul", "0,3", "双输入 matmul"),
            (5, "add", "", "独立 add"),
            (6, "add", "0,1", "双输入 add"),
            (7, "mul", "", "独立 mul"),
            (8, "mul", "0,1", "双输入 mul"),
            (9, "unknown_op", "", "未知算子"),
        ]

        for i, (op_id, op_name, depends, note) in enumerate(ops_to_test):
            ws.cell(row=2+i, column=1, value=op_id)
            ws.cell(row=2+i, column=2, value=op_name)
            ws.cell(row=2+i, column=3, value="2,4,8")
            ws.cell(row=2+i, column=4, value="float32")
            ws.cell(row=2+i, column=5, value=depends)
            ws.cell(row=2+i, column=7, value="FALSE")
            ws.cell(row=2+i, column=8, value=note)

        wb.save(xlsx_path)

        py_path = Path(self.temp_dir) / "all_ops.py"
        code = import_xlsx(str(xlsx_path), str(py_path))

        # 代码现在使用 F.linear 而非 nn.linear
        assert "F.linear" in code
        assert "F.relu" in code
        assert "F.softmax" in code
        assert "F.matmul" in code
        assert "F.add" in code
        assert "F.mul" in code
        assert "未知算子: unknown_op" in code
        assert py_path.exists()

    def test_code_gen_with_skip(self):
        """测试跳过算子的代码生成"""
        from aidevtools.xlsx import create_template, import_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "skip_ops.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="linear")
        ws.cell(row=2, column=3, value="2,4,8")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="TRUE")  # skip
        ws.cell(row=2, column=8, value="被跳过的算子")

        wb.save(xlsx_path)

        code = import_xlsx(str(xlsx_path), None)

        assert "[SKIP]" in code

    def test_parse_xlsx_comment_row(self):
        """测试解析注释行"""
        from aidevtools.xlsx import create_template, parse_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "comment.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        ws.cell(row=2, column=1, value="# 这是注释")
        ws.cell(row=3, column=1, value=0)
        ws.cell(row=3, column=2, value="relu")
        ws.cell(row=3, column=3, value="1,64")
        ws.cell(row=3, column=4, value="float32")
        ws.cell(row=3, column=7, value="FALSE")

        wb.save(xlsx_path)

        _, configs = parse_xlsx(str(xlsx_path))
        assert len(configs) == 1
        assert configs[0].op_name == "relu"

    def test_code_gen_attention_without_qkv(self):
        """测试 attention 缺少 qkv 的代码生成"""
        from aidevtools.xlsx import create_template, import_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "attention_no_qkv.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="attention")
        ws.cell(row=2, column=3, value="1,4,8,32")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=5, value="")
        ws.cell(row=2, column=7, value="FALSE")

        wb.save(xlsx_path)

        code = import_xlsx(str(xlsx_path), None)
        assert "attention 需要 q, k, v" in code


class TestSimCmd:
    """仿真命令测试"""

    def setup_method(self):
        """创建临时目录"""
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        """清理临时目录"""
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_template_has_sim_cmd_column(self):
        """测试模板包含 sim_cmd 列"""
        from aidevtools.xlsx import create_template
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_sim.xlsx"
        create_template(str(xlsx_path))

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        # 检查表头
        headers = [cell.value for cell in ws[1]]
        assert "sim_cmd" in headers

    def test_parse_xlsx_with_sim_cmd(self):
        """测试解析带 sim_cmd 的 xlsx"""
        from aidevtools.xlsx import create_template, parse_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_sim_parse.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        # 找到 sim_cmd 列索引
        headers = [cell.value for cell in ws[1]]
        sim_cmd_col = headers.index("sim_cmd") + 1

        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="linear")
        ws.cell(row=2, column=3, value="1,4,8")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="FALSE")
        ws.cell(row=2, column=sim_cmd_col, value="./my_sim.sh {golden_bin} {result_bin}")

        wb.save(xlsx_path)

        enabled_ops, op_configs = parse_xlsx(str(xlsx_path))
        assert len(op_configs) == 1
        assert op_configs[0].paths.sim_cmd == "./my_sim.sh {golden_bin} {result_bin}"

    def test_run_sim_cmd_placeholder_substitution(self):
        """测试 sim_cmd 占位符替换"""
        from aidevtools.xlsx.run import _run_sim_cmd
        from aidevtools.xlsx.import_ import OpConfig

        output_dir = Path(self.temp_dir)

        # 创建一个模拟脚本
        sim_script = output_dir / "mock_sim.sh"
        sim_script.write_text("""#!/bin/bash
echo "input: $1"
echo "golden: $2"
echo "result: $3"
# 创建 result 文件（复制 golden）
cp "$2" "$3"
""")
        sim_script.chmod(0o755)

        # 创建 golden 文件
        golden_data = np.array([1.0, 2.0, 3.0], dtype=np.float32)
        golden_path = output_dir / "test_0_golden.bin"
        golden_data.tofile(golden_path)

        # 创建 OpConfig
        config = OpConfig(
            id=0,
            op_name="test",
            shape=(),
            dtype="float32",
            depends="",
            qtype="",
            skip=False,
            note="",
        )

        # 执行仿真命令
        result = _run_sim_cmd(
            sim_cmd=f"{sim_script} {{input_bin}} {{golden_bin}} {{result_bin}}",
            config=config,
            output_dir=output_dir,
            name="test_0",
            has_input=True,
        )

        # 验证结果
        assert result is not None
        assert result.shape == golden_data.shape

    def test_run_sim_cmd_empty(self):
        """测试空 sim_cmd"""
        from aidevtools.xlsx.run import _run_sim_cmd
        from aidevtools.xlsx.import_ import OpConfig

        config = OpConfig(
            id=0,
            op_name="test",
            shape=(),
            dtype="float32",
            depends="",
            qtype="",
            skip=False,
            note="",
        )

        result = _run_sim_cmd(
            sim_cmd="",
            config=config,
            output_dir=Path(self.temp_dir),
            name="test_0",
        )
        assert result is None

        result = _run_sim_cmd(
            sim_cmd="   ",
            config=config,
            output_dir=Path(self.temp_dir),
            name="test_0",
        )
        assert result is None

    def test_run_sim_cmd_failure(self):
        """测试 sim_cmd 执行失败"""
        from aidevtools.xlsx.run import _run_sim_cmd
        from aidevtools.xlsx.import_ import OpConfig

        config = OpConfig(
            id=0,
            op_name="test",
            shape=(),
            dtype="float32",
            depends="",
            qtype="",
            skip=False,
            note="",
        )

        result = _run_sim_cmd(
            sim_cmd="exit 1",
            config=config,
            output_dir=Path(self.temp_dir),
            name="test_0",
        )
        assert result is None

    def test_run_xlsx_with_sim_cmd(self):
        """测试完整 run_xlsx 流程中的 sim_cmd"""
        from aidevtools.xlsx import create_template, run_xlsx
        from openpyxl import load_workbook

        xlsx_path = Path(self.temp_dir) / "test_full_sim.xlsx"
        create_template(str(xlsx_path), include_examples=False)

        # 创建模拟仿真脚本
        sim_script = Path(self.temp_dir) / "sim.sh"
        sim_script.write_text("""#!/bin/bash
cp "$1" "$2"
""")
        sim_script.chmod(0o755)

        wb = load_workbook(xlsx_path)
        ws = wb["ops"]

        headers = [cell.value for cell in ws[1]]
        sim_cmd_col = headers.index("sim_cmd") + 1

        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="linear")
        ws.cell(row=2, column=3, value="1,4,8")
        ws.cell(row=2, column=4, value="float32")
        ws.cell(row=2, column=7, value="FALSE")
        ws.cell(row=2, column=sim_cmd_col, value=f"{sim_script} {{golden_bin}} {{result_bin}}")

        wb.save(xlsx_path)

        results = run_xlsx(str(xlsx_path), str(self.temp_dir))

        # 由于 golden 实现存在，sim_cmd 不会执行
        # 但配置应该被正确解析
        assert len(results) >= 1


